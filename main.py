from pandas import read_excel, ExcelFile
from tkinter import filedialog as fd
from datetime import datetime, date
from copy import copy
import tkinter as tk
import traceback
import openpyxl
import time
import os
import re

sap_file = ''
bank_file = ''
output_folder = ''


def read_sap_excel_file(excel_path):
    columns = ['Nº documento', 'Fecha de documento', 'Asignación', 'Importe en moneda local']
    excel_file = ExcelFile(excel_path)
    sheet_name = [sn for sn in excel_file.sheet_names if 'BBVA 0829' in sn][0]
    excel_df = read_excel(excel_path, sheet_name=sheet_name, dtype=str, engine='openpyxl').round(2)
    return excel_df[list(columns)].fillna('').rename(columns={'Nº documento': 'N. Doc', 'Fecha de documento': 'Date', 'Asignación': 'Assignation',
                                                    'Importe en moneda local': 'Ammount'}).to_dict(orient='records')


def read_bank_excel_file(excel_path):
    columns = ['Concepto', 'Retiro', 'Fecha Operación', 'Cuenta', 'Deposito']
    excel_file = ExcelFile(excel_path)
    sheet_name = [sn for sn in excel_file.sheet_names if '0829' in sn][0]
    excel_df = read_excel(excel_path, sheet_name=sheet_name, dtype=str, engine='openpyxl').round(2)
    return excel_df[list(columns)].fillna('').rename(columns={'Concepto': 'Concept', 'Retiro': 'Withdrawal', 'Fecha Operación': 'Date',
                                                    'Cuenta': 'Account', 'Deposito': 'Deposit'}).to_dict(orient='records')


def set_movement_for_bank_records(bank_records):
    for record in bank_records:
        record['Movement'] = 'Charge' if record['Withdrawal'] != '' else 'Payment'


def set_movement_for_sap_records(bank_records):
    for record in bank_records:
        record['Movement'] = 'Charge' if float(record['Ammount']) > 0 else 'Payment'


def get_bank_records(bank_file):
    bank_records = read_bank_excel_file(bank_file)
    set_movement_for_bank_records(bank_records)
    bank_records = list(filter(lambda record: record['Account'] != '' and record['Date'] != '', bank_records))
    for bank_record in bank_records:
        if bank_record['Deposit'] != '':
            bank_record['Deposit'] = "{:.2f}".format(float(bank_record['Deposit']))
        if bank_record['Withdrawal'] != '':
            bank_record['Withdrawal'] = "{:.2f}".format(float(bank_record['Withdrawal']))

    return bank_records


def get_sap_records(sap_file):
    sap_records = read_sap_excel_file(sap_file)
    set_movement_for_sap_records(sap_records)
    sap_records = list(filter(lambda record: record['Assignation'] != '' and record['Date'] != '', sap_records))
    for sap_record in sap_records:
        if float(sap_record['Ammount']) < 0:
            sap_record['Ammount'] = "{:.2f}".format(float(sap_record['Ammount']) * (-1))
        else:
            sap_record['Ammount'] = "{:.2f}".format(float(sap_record['Ammount']))

    return sap_records


def get_payments_from_the_bank_not_reciprocated_by_us(bank_records, sap_records):
    payments_from_the_bank_not_reciprocated_by_us = []
    bank_payments = [bank_record for bank_record in bank_records if bank_record['Movement'] == 'Payment']
    sap_charges = [sap_record for sap_record in sap_records if sap_record['Movement'] == 'Charge']

    for bank_record in bank_payments:
        sap_record = next((record for record in sap_charges if record['Date'] == bank_record['Date'] and bank_record['Deposit'] == record['Ammount']), None)
        if sap_record is None:
            payments_from_the_bank_not_reciprocated_by_us.append(bank_record)

    return payments_from_the_bank_not_reciprocated_by_us


def get_charges_from_the_bank_not_reciprocated_by_us(bank_records, sap_records):
    charges_from_the_bank_not_reciprocated_by_us = []
    bank_charges = [bank_record for bank_record in bank_records if bank_record['Movement'] == 'Charge']
    sap_payments = [sap_record for sap_record in sap_records if sap_record['Movement'] == 'Payment']

    for bank_record in bank_charges:
        sap_record = next((record for record in sap_payments if record['Date'] == bank_record['Date'] and bank_record['Withdrawal'] == record['Ammount']), None)
        if sap_record is None:
            charges_from_the_bank_not_reciprocated_by_us.append(bank_record)

    return charges_from_the_bank_not_reciprocated_by_us


def get_our_payments_not_reciprocated_by_the_bank(bank_records, sap_records):
    our_payments_not_reciprocated_by_the_bank = []
    bank_charges = [bank_record for bank_record in bank_records if bank_record['Movement'] == 'Charge']
    sap_payments = [sap_record for sap_record in sap_records if sap_record['Movement'] == 'Payment']

    for sap_record in sap_payments:
        bank_record = next((record for record in bank_charges if record['Date'] == sap_record['Date'] and sap_record['Ammount'] == record['Withdrawal']), None)
        if bank_record is None:
            our_payments_not_reciprocated_by_the_bank.append(sap_record)

    return our_payments_not_reciprocated_by_the_bank


def get_our_charges_not_reciprocated_by_the_bank(bank_records, sap_records):
    our_charges_not_reciprocated_by_the_bank = []
    bank_payments = [bank_record for bank_record in bank_records if bank_record['Movement'] == 'Payment']
    sap_charges = [sap_record for sap_record in sap_records if sap_record['Movement'] == 'Charge']

    for sap_record in sap_charges:
        bank_record = next((record for record in bank_payments if record['Date'] == sap_record['Date'] and sap_record['Ammount'] == record['Deposit']), None)
        if bank_record is None:
            our_charges_not_reciprocated_by_the_bank.append(sap_record)

    return our_charges_not_reciprocated_by_the_bank


def copy_rows(sheet, row_idx):
    row = sheet[row_idx]
    sheet.insert_rows(row_idx)
    for cell in row:
        new_cell = sheet.cell(row=row_idx, column=cell.col_idx)
        if cell.has_style:
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)

def run(sap_file, bank_file, output_folder):
    start_time = time.time()

    if sap_file == '' or bank_file == '' or output_folder == '':
        tk.messagebox.showwarning(title='Requerimientos incompletos', message=f'Algunos campos no fueron llenados. Asegurate de elegir todo lo que se te pide.')
        return

    bank_reconciliation_template = 'templates\\template.xlsx'

    try:
        bank_records = get_bank_records(bank_file)
    except KeyError as key_error:
        print(traceback.format_exc())
        colums_str = re.findall('\[.*?\]', str(key_error))[0].replace('[', '').replace(']', '').replace(', ', '\n')
        tk.messagebox.showwarning(title='Algunas columnas no se encontraron', message=f'El sistema no pudo encontrar las siguientes columnas en el archivo del banco:\n\n{colums_str}')
        return
    except IndexError:
        print(traceback.format_exc())
        tk.messagebox.showwarning(title='Hoja no encontrada', message=f'Asegurate de que tu archivo del banco contiene una hoja con el texto "0829".')
        return
    except Exception:
        print(traceback.format_exc())
        tk.messagebox.showerror(title='Error inesperado', message=f'Ha ocurrido un error inesperado al leer el achivo del banco. Para recibir soporte, contacta al desarrollador enviando los archivos de entrada y una captura de este error:\n\n{traceback.format_exc()}')
        return

    try:
        sap_records = get_sap_records(sap_file)
    except KeyError as key_error:
        print(traceback.format_exc())
        colums_str = re.findall('\[.*?\]', str(key_error))[0].replace('[', '').replace(']', '').replace(', ', '\n')
        tk.messagebox.showwarning(title='Algunas columnas no se encontraron', message=f'El sistema no pudo encontrar las siguientes columnas en el archivo del SAP:\n\n{colums_str}')
        return
    except IndexError:
        print(traceback.format_exc())
        tk.messagebox.showwarning(title='Hoja no encontrada', message=f'Asegurate de que tu archivo del SAP contiene una hoja con el texto "BBVA 0829".')
        return
    except Exception:
        print(traceback.format_exc())
        tk.messagebox.showerror(title='Error inesperado', message=f'Ha ocurrido un error inesperado al leer el achivo del SAP. Para recibir soporte, contacta al desarrollador enviando los archivos de entrada y una captura de este error:\n\n{traceback.format_exc()}')
    
    try:
        our_charges_not_reciprocated_by_the_bank = get_our_charges_not_reciprocated_by_the_bank(bank_records, sap_records)
        our_payments_not_reciprocated_by_the_bank = get_our_payments_not_reciprocated_by_the_bank(bank_records, sap_records)
        payments_from_the_bank_not_reciprocated_by_us = get_payments_from_the_bank_not_reciprocated_by_us(bank_records, sap_records)
        charges_from_the_bank_not_reciprocated_by_us = get_charges_from_the_bank_not_reciprocated_by_us(bank_records, sap_records)

        reconciliation_file = openpyxl.load_workbook(bank_reconciliation_template)
        sheet = reconciliation_file['BBVA 0829']

        row_idx = 12
        for record in our_charges_not_reciprocated_by_the_bank:
            copy_rows(sheet, row_idx)
            sheet.cell(row=row_idx, column=4).value = record['N. Doc']
            sheet.cell(row=row_idx, column=5).value = datetime.strptime(record['Date'], '%Y-%m-%d %H:%M:%S').strftime('%d/%m/%Y')
            sheet.cell(row=row_idx, column=6).value = record['Assignation']
            sheet.cell(row=row_idx, column=7).value = float(record['Ammount'])

        row_idx += len(our_charges_not_reciprocated_by_the_bank) + 2
        for record in charges_from_the_bank_not_reciprocated_by_us:
            copy_rows(sheet, row_idx)
            sheet.cell(row=row_idx, column=4).value = record['Account']
            sheet.cell(row=row_idx, column=5).value = datetime.strptime(record['Date'], '%Y-%m-%d %H:%M:%S').strftime('%d/%m/%Y')
            sheet.cell(row=row_idx, column=6).value = record['Concept']
            sheet.cell(row=row_idx, column=7).value = float(record['Withdrawal'])

        row_idx += len(charges_from_the_bank_not_reciprocated_by_us) + 2
        for record in our_payments_not_reciprocated_by_the_bank:
            copy_rows(sheet, row_idx)
            sheet.cell(row=row_idx, column=4).value = record['N. Doc']
            sheet.cell(row=row_idx, column=5).value = datetime.strptime(record['Date'], '%Y-%m-%d %H:%M:%S').strftime('%d/%m/%Y')
            sheet.cell(row=row_idx, column=6).value = record['Assignation']
            sheet.cell(row=row_idx, column=7).value = float(record['Ammount'])

        row_idx += len(our_payments_not_reciprocated_by_the_bank) + 2
        for record in payments_from_the_bank_not_reciprocated_by_us:
            copy_rows(sheet, row_idx)
            sheet.cell(row=row_idx, column=4).value = record['Account']
            sheet.cell(row=row_idx, column=5).value = datetime.strptime(record['Date'], '%Y-%m-%d %H:%M:%S').strftime('%d/%m/%Y')
            sheet.cell(row=row_idx, column=6).value = record['Concept']
            sheet.cell(row=row_idx, column=7).value = float(record['Deposit'])

        now = date.today()
        version = 0
        output_file = os.path.join(output_folder, f'Conciliación {now.strftime("%d %B %Y")}.xlsx')
        while os.path.exists(output_file):
            version += 1
            output_file = os.path.join(output_folder, f'Conciliación {now.strftime("%d %B %Y")} ({str(version)}).xlsx')

        reconciliation_file.save(output_file)
        reconciliation_file.close()
        tk.messagebox.showinfo(title='Proceso terminado con éxito', message=f'El archivo de conciliación se ha generado satisfactoriamente en {"{:.2f}".format(time.time() - start_time)} segundos.')
    except FileNotFoundError as file_not_found_error:
        print(traceback.format_exc())
        tk.messagebox.showwarning(title='Template de conciliación no encontrado', message=f'El template de conciliación no se encontró. Asegurate de haber instalado correctamente el sistema. Contacta al desarrollador si el problema persiste.')
    except KeyError as key_error:
        print(traceback.format_exc())
        if 'Worksheet' in str(key_error):
            tk.messagebox.showwarning(title='Hoja no encontrada', message=f'Asegurate de que el template de conciliación está instalado correctamente y no ha sufrido ningún cambio.')
        else:
            tk.messagebox.showerror(title='Error inesperado', message=f'Ha ocurrido un error inesperado al generar la conciliación. Para recibir soporte, contacta al desarrollador enviando los archivos de entrada.')
    except Exception:
        print(traceback.format_exc())
        tk.messagebox.showerror(title='Error inesperado', message=f'Ha ocurrido un error inesperado al generar la conciliación. Para recibir soporte, contacta al desarrollador enviando los archivos de entrada y una captura de este error:\n\n{traceback.format_exc()}')


def select_files(entry, canvas, x, y, file):
    filename = fd.askopenfilename(
        title='Open files',
        initialdir='/',
        filetypes=[("Excel files", "*.xlsx")]
    )

    if file == 'sap':
        global sap_file
        sap_file = filename
    elif file == 'bank':
        global bank_file
        bank_file = filename

    entry['state'] = 'normal'
    entry.delete(0, tk.END)
    entry.insert(0, filename)
    entry.pack()
    entry['state'] = 'disabled'
    canvas.create_window(x, y, window=entry)


def select_folder(entry, canvas, x, y):
    foldername = fd.askdirectory(
        title='Open files',
        initialdir='/'
    )

    global output_folder
    output_folder = foldername

    entry['state'] = 'normal'
    entry.delete(0, tk.END)
    entry.insert(0, foldername)
    entry.pack()
    entry['state'] = 'disabled'
    canvas.create_window(x, y, window=entry)


if __name__ == '__main__':
    # Define coordenates of elements
    sap_label_x = 75
    sap_label_y = 40
    sap_entry_x = 215
    sap_entry_y = 60
    sap_button_x = 465
    sap_button_y = 60
    bank_label_x = 83
    bank_label_y = 100
    bank_entry_x = 215
    bank_entry_y = 120
    bank_button_x = 465
    bank_button_y = 120
    folder_label_x = 77
    folder_label_y = 160
    folder_entry_x = 215
    folder_entry_y = 180
    folder_button_x = 465
    folder_button_y = 180

    # Configure window
    root = tk.Tk()
    root.iconbitmap('img\\logo.ico')
    root.title('Conciliaciones bancarias')
    root.resizable(False, False)
    canvas = tk.Canvas(root, width = 550, height = 300, bg='#fff')
    canvas.pack()

    # Configure SAP elements
    sap_label = tk.Label(root, text='Archivo de SAP', bg='#fff', fg='#000', font='Helvetica 8 bold')
    canvas.create_window(sap_label_x, sap_label_y, window=sap_label)

    sap_entry = tk.Entry(root, width=60, state='disabled', disabledbackground='#ddd', disabledforeground='#3d5277', font='Helvetica 8 bold')
    canvas.create_window(sap_entry_x, sap_entry_y, window=sap_entry)

    sap_button = tk.Button(text='Seleccionar archivo', bg='#3d5277', font='Helvetica 8 bold', fg='#fff', command=lambda:select_files(sap_entry, canvas, sap_entry_x, sap_entry_y, 'sap'))
    canvas.create_window(sap_button_x, sap_button_y, window=sap_button)

    # Configure bank elements
    bank_label = tk.Label(root, text='Archivo del banco', bg='#fff', fg='#000', font='Helvetica 8 bold')
    canvas.create_window(bank_label_x, bank_label_y, window=bank_label)

    bank_entry = tk.Entry(root, width=60, state='disabled', disabledbackground='#ddd', font='Helvetica 8 bold', disabledforeground='#3d5277')
    canvas.create_window(bank_entry_x, bank_entry_y, window=bank_entry)

    bank_button = tk.Button(text='Seleccionar archivo', bg='#3d5277', font='Helvetica 8 bold', fg='#fff', command=lambda:select_files(bank_entry, canvas, bank_entry_x, bank_entry_y, 'bank'))
    canvas.create_window(bank_button_x, bank_button_y, window=bank_button)

    # Configure output folder elements
    folder_label = tk.Label(root, text='Folder de salida', bg='#fff', fg='#000', font='Helvetica 8 bold')
    canvas.create_window(folder_label_x, folder_label_y, window=folder_label)

    folder_entry = tk.Entry(root, width=60, state='disabled', disabledbackground='#ddd', font='Helvetica 8 bold', disabledforeground='#3d5277')
    canvas.create_window(folder_entry_x, folder_entry_y, window=folder_entry)

    folder_button = tk.Button(text='Seleccionar carpeta', bg='#3d5277', fg='#fff', font='Helvetica 8 bold', command=lambda:select_folder(folder_entry, canvas, folder_entry_x, folder_entry_y))
    canvas.create_window(folder_button_x, folder_button_y, window=folder_button)
    
    # Configure submit button
    submit_button = tk.Button(text='Generar conciliación', width=20, height=2, bg='#2b3b58', fg='#fff', font='Helvetica 9 bold', command=lambda:run(sap_file, bank_file, output_folder))
    canvas.create_window(275, 245, window=submit_button)

    # Show window
    root.mainloop()